from sqlalchemy.orm import Session

import openpyxl as xl
from io import BytesIO
from src.infrastructure.database.orm_models import PlanEstudios
from src.infrastructure.api.schemas.planes_estudios_schema import PlanEstudiosCreate, PlanEstudiosUpdate

def crear_nuevo_plan_estudios(db: Session, plan_data: PlanEstudiosCreate):
    try:
        nuevo_plan = PlanEstudios(**plan_data.model_dump())
    except Exception as e:
        raise ValueError(f"Error al crear el plan de estudios: {str(e)}")
    
    db.add(nuevo_plan)
    db.commit()
    db.refresh(nuevo_plan)
    
    return nuevo_plan

def obtener_todos_los_planes_estudios(db: Session):
    return db.query(PlanEstudios).all()

def obtener_plan_estudios_por_id(db: Session, plan_id: int):
    return db.query(PlanEstudios).filter(PlanEstudios.id == plan_id).first()

def actualizar_plan_estudios(db: Session, plan_id: int, plan_data: PlanEstudiosUpdate):
    plan = db.query(PlanEstudios).filter(PlanEstudios.id == plan_id).first()
    if not plan:
        raise ValueError("Plan de estudios no encontrado")
    
    plan_data_dict = plan_data.model_dump(exclude_unset=True)
    
    for key, value in plan_data_dict.items():
        setattr(plan, key, value)
    
    db.commit()
    db.refresh(plan)
    
    return plan

def eliminar_plan_estudios(db: Session, plan_id: int):
    from src.infrastructure.database.orm_models import PlanEstudios
    plan = db.query(PlanEstudios).filter(PlanEstudios.id == plan_id).first()
    if not plan:
        raise ValueError("Plan de estudios no encontrado")
    
    # 1. Obtener todas las materias del plan y eliminar sus asignaciones y horarios
    for materia in plan.materias:
        for asignacion in materia.asignaciones:
            for horario in asignacion.horarios:
                db.delete(horario)
            db.delete(asignacion)
        db.delete(materia)
        
    # 2. Obtener todos los grupos del plan y eliminar sus asignaciones y horarios
    for grupo in plan.grupos_abiertos:
        for asignacion in grupo.asignaciones:
            for horario in asignacion.horarios:
                db.delete(horario)
            db.delete(asignacion)
        db.delete(grupo)
        
    # 3. Eliminar el plan de estudios
    db.delete(plan)
    db.commit()
    return True

async def importar_plan_estudios(db: Session, byte_object: BytesIO):
    try:
        from src.infrastructure.database.orm_models import ProgramaEducativo
        wb = xl.load_workbook(byte_object)
        sheet = wb["plan_estudios"]
        objects = []
        headers = [cell.value for cell in sheet[1]]

        # Construir mapeo de ID de Excel a Clave/Nombre del Programa Educativo si existe la pestaña en el archivo
        excel_prog_map = {}
        if "programas_educativos" in wb.sheetnames:
            prog_sheet = wb["programas_educativos"]
            prog_headers = [c.value for c in prog_sheet[1]]
            for r in prog_sheet.iter_rows(min_row=2, values_only=True):
                r_data = {k: v for k, v in zip(prog_headers, r)}
                ex_id = r_data.get("id")
                clave = r_data.get("clave")
                nombre = r_data.get("nombre")
                if ex_id is not None:
                    excel_prog_map[int(ex_id)] = (clave, nombre)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {key: value for key, value in zip(headers, row)}
            if not any(value is not None and str(value).strip() != "" for value in row_data.values()):
                continue

            prog_id = None
            prog_id_raw = row_data.get("programa_educativo_id")
            if prog_id_raw is not None and str(prog_id_raw).strip() != "":
                # 1. Intentar mapear usando nuestro mapa de Excel
                try:
                    ex_id = int(prog_id_raw)
                    if ex_id in excel_prog_map:
                        clave, nombre = excel_prog_map[ex_id]
                        if clave:
                            prog = db.query(ProgramaEducativo).filter(ProgramaEducativo.clave == clave).first()
                            if prog:
                                prog_id = prog.id
                        if not prog_id and nombre:
                            prog = db.query(ProgramaEducativo).filter(ProgramaEducativo.nombre == nombre).first()
                            if prog:
                                prog_id = prog.id
                except ValueError:
                    pass

                # 2. Si no se resolvió, ver si el ID existe directo en BD
                if not prog_id:
                    try:
                        p_id = int(prog_id_raw)
                        exists = db.query(ProgramaEducativo).filter(ProgramaEducativo.id == p_id).first()
                        if exists:
                            prog_id = exists.id
                    except ValueError:
                        pass
            
            # 3. Si sigue sin resolverse, buscar por claves/nombres en otros campos del Excel
            if not prog_id:
                prog_clave = str(row_data.get("programa_educativo_clave") or row_data.get("programa_clave") or "").strip()
                if prog_clave:
                    prog = db.query(ProgramaEducativo).filter(ProgramaEducativo.clave == prog_clave).first()
                    if prog:
                        prog_id = prog.id
                
                if not prog_id:
                    prog_name = str(row_data.get("programa_educativo") or row_data.get("programa") or "").strip()
                    if prog_name:
                        prog = db.query(ProgramaEducativo).filter(ProgramaEducativo.nombre == prog_name).first()
                        if prog:
                            prog_id = prog.id

            # 4. Si aún no se resolvió, intentar deducir a partir del nombre del plan (si contiene la clave del programa)
            if not prog_id:
                plan_name = str(row_data.get("nombre") or "").strip()
                all_progs = db.query(ProgramaEducativo).all()
                for p in all_progs:
                    if p.clave and p.clave in plan_name:
                        prog_id = p.id
                        break

            if not prog_id:
                raise ValueError(f"No se pudo resolver el Programa Educativo para la fila con nombre: '{row_data.get('nombre')}'")

            plan_data = PlanEstudiosCreate(
                nombre=str(row_data.get("nombre") or "").strip(),
                programa_educativo_id=prog_id,
                vigente=row_data.get("vigente") if row_data.get("vigente") is not None else True,
                tipo_periodo=row_data.get("tipo_periodo")
            )
            nuevo_plan = crear_nuevo_plan_estudios(db, plan_data)
            objects.append(nuevo_plan)

        return objects
    except Exception as e:
        raise ValueError(f"Error al crear el plan de estudios: {str(e)}")
    
async def exportar_plan_estudios(db: Session):
    planes = db.query(PlanEstudios).all()
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "planes_estudio"

    headers = ["id", "nombre", "programa_educativo_id", "vigente", "tipo_periodo"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    for row_idx, plan in enumerate(planes, start=2):
        sheet.cell(row=row_idx, column=1, value=plan.id)
        sheet.cell(row=row_idx, column=2, value=plan.nombre)
        sheet.cell(row=row_idx, column=3, value=plan.programa_educativo_id)
        sheet.cell(row=row_idx, column=4, value=plan.vigente)
        sheet.cell(row=row_idx, column=5, value=plan.tipo_periodo.value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer