from io import BytesIO

import openpyxl as xl
from sqlalchemy.orm import Session

from src.infrastructure.api.schemas.areas_schema import AreaConocimientoCreate, AreaConocimientoUpdate
from src.infrastructure.database.orm_models import AreaConocimiento


def crear_area(db: Session, area_data: AreaConocimientoCreate):
    nueva_area = AreaConocimiento(**area_data.model_dump())

    db.add(nueva_area)
    db.commit()
    db.refresh(nueva_area)

    return nueva_area


def obtener_areas(db: Session):
    return db.query(AreaConocimiento).all()


def obtener_area_por_id(db: Session, area_id: int):
    return db.query(AreaConocimiento).filter(AreaConocimiento.id == area_id).first()


def actualizar_area(db: Session, area_id: int, area_data: AreaConocimientoUpdate):
    area = db.query(AreaConocimiento).filter(AreaConocimiento.id == area_id).first()
    if not area:
        raise ValueError("Área de conocimiento no encontrada")

    area_data_dict = area_data.model_dump(exclude_unset=True)

    for key, value in area_data_dict.items():
        setattr(area, key, value)

    db.commit()
    db.refresh(area)

    return area


def eliminar_area(db: Session, area_id: int):
    area = db.query(AreaConocimiento).filter(AreaConocimiento.id == area_id).first()
    if not area:
        raise ValueError("Área de conocimiento no encontrada")

    db.delete(area)
    db.commit()
    return True


async def importar_areas(db: Session, byte_object: BytesIO):
    try:
        byte_object.seek(0)
        wb = xl.load_workbook(byte_object)
        sheet = wb["areas_conocimiento"]
        objects = []
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {key: value for key, value in zip(headers, row)}
            if not any(value is not None and str(value).strip() != "" for value in row_data.values()):
                continue

            area_data = AreaConocimientoCreate(
                nombre=str(row_data.get("nombre") or "").strip(),
                descripcion=str(row_data.get("descripcion") or "").strip() if row_data.get("descripcion") not in (None, "") else None,
            )
            objects.append(crear_area(db, area_data))

        return objects
    except Exception as e:
        raise ValueError(f"Error al importar las áreas de conocimiento: {str(e)}")


async def exportar_areas(areas: list):
    wb = xl.Workbook()
    sheet = wb.active
    
    if sheet is None:
        raise ValueError("No se pudo crear la hoja de cálculo para exportar las áreas de conocimiento.")
    
    sheet.title = "areas_conocimiento" 

    headers = ["id", "nombre", "descripcion"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    for row_idx, area in enumerate(areas, start=2):
        sheet.cell(row=row_idx, column=1, value=area.id)
        sheet.cell(row=row_idx, column=2, value=area.nombre)
        sheet.cell(row=row_idx, column=3, value=area.descripcion)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
    