from io import BytesIO

import openpyxl as xl
from sqlalchemy.orm import Session

from src.infrastructure.api.schemas.programas_schema import ProgramaEducativoCreate, ProgramaEducativoUpdate
from src.infrastructure.database.orm_models import ProgramaEducativo


def crear_nuevo_programa(db: Session, programa_data: ProgramaEducativoCreate):
    if db.query(ProgramaEducativo).filter_by(clave=programa_data.clave.upper()).first():
        raise ValueError(f"Ya existe un programa educativo con la clave '{programa_data.clave.upper()}'")

    try:
        data_dict = programa_data.model_dump()
        data_dict['clave'] = data_dict['clave'].upper()
        nuevo_programa = ProgramaEducativo(**data_dict)
    except Exception as e:
        raise ValueError(f"Error al crear el programa educativo: {str(e)}")

    db.add(nuevo_programa)
    db.commit()
    db.refresh(nuevo_programa)

    return nuevo_programa


def obtener_todos_los_programas(db: Session, unidad_id: int | None = None):
    q = db.query(ProgramaEducativo)
    if unidad_id is not None:
        q = q.filter(ProgramaEducativo.unidad_academica_id == unidad_id)
    return q.all()


def obtener_programa_por_id(db: Session, programa_id: int):
    programa = db.query(ProgramaEducativo).filter(ProgramaEducativo.id == programa_id).first()

    if not programa:
        raise ValueError(f"No se encontró un programa educativo con el ID {programa_id}")

    return programa


def actualizar_programa(db: Session, programa_id: int, programa_data: ProgramaEducativoUpdate):
    db_programa = db.query(ProgramaEducativo).filter(ProgramaEducativo.id == programa_id).first()

    if not db_programa:
        return None

    datos_actualizar = programa_data.model_dump(exclude_unset=True)

    if "clave" in datos_actualizar:
        datos_actualizar["clave"] = datos_actualizar["clave"].upper()

    for clave, valor in datos_actualizar.items():
        setattr(db_programa, clave, valor)

    db.commit()
    db.refresh(db_programa)
    return db_programa


def eliminar_programa(db: Session, programa_id: int):
    from src.infrastructure.database.orm_models import PlanEstudios
    from src.application.use_cases.planes_estudios_service import eliminar_plan_estudios
    
    db_programa = db.query(ProgramaEducativo).filter(ProgramaEducativo.id == programa_id).first()

    if not db_programa:
        return False

    # 1. Obtener todos los planes asociados a este programa y eliminarlos en cascada
    planes = db.query(PlanEstudios).filter(PlanEstudios.programa_educativo_id == programa_id).all()
    for plan in planes:
        eliminar_plan_estudios(db, plan.id) #type: ignore

    # 2. Eliminar el programa
    db.delete(db_programa)
    db.commit()
    return True


async def importar_programas(db: Session, byte_object: BytesIO):
    try:
        byte_object.seek(0)
        wb = xl.load_workbook(byte_object)
        sheet = wb["programas_educativos"]
        objects = []
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {key: value for key, value in zip(headers, row)}
            if not any(value is not None and str(value).strip() != "" for value in row_data.values()):
                continue

            programa_data = ProgramaEducativoCreate(
                nombre=str(row_data.get("nombre") or "").strip(),
                clave=str(row_data.get("clave") or "").strip(),
                activo=bool(row_data.get("activo", True)) if row_data.get("activo") is not None else True,
                nivel=str(row_data.get("nivel") or "LICENCIATURA").strip().upper()
            )
            objects.append(crear_nuevo_programa(db, programa_data))

        return objects
    except Exception as e:
        raise ValueError(f"Error al importar los programas educativos: {str(e)}")


async def exportar_programas(programas: list):
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "programas_educativos" #type: ignore

    headers = ["id", "nombre", "clave", "activo", "nivel"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header) #type: ignore

    for row_idx, programa in enumerate(programas, start=2):
        sheet.cell(row=row_idx, column=1, value=programa.id) #type: ignore
        sheet.cell(row=row_idx, column=2, value=programa.nombre) #type: ignore
        sheet.cell(row=row_idx, column=3, value=programa.clave) #type: ignore
        sheet.cell(row=row_idx, column=4, value=programa.activo) #type: ignore
        sheet.cell(row=row_idx, column=5, value=programa.nivel) #type: ignore

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer