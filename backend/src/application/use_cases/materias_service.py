from io import BytesIO

import openpyxl as xl
from sqlalchemy.orm import Session

from src.infrastructure.api.schemas.materias_schema import MateriaCreate, MateriaUpdate
from src.infrastructure.database.orm_models import Materia


def crear_nueva_materia(db: Session, materia_data: MateriaCreate):
    # Evitar duplicados (mismo nombre, plan, periodo y hsm)
    existente = db.query(Materia).filter(
        Materia.nombre_asignatura == materia_data.nombre_asignatura,
        Materia.plan_estudios_id == materia_data.plan_estudios_id,
        Materia.numero_periodo == materia_data.numero_periodo,
        Materia.hsm == materia_data.hsm
    ).first()
    if existente:
        return existente

    try:
        nueva_materia = Materia(**materia_data.model_dump())
    except Exception as e:
        raise ValueError(f"Error al crear la materia: {str(e)}")
    db.add(nueva_materia)
    db.commit()
    db.refresh(nueva_materia)
    return nueva_materia


def obtener_todas_las_materias(db: Session):
    return db.query(Materia).all()


def obtener_materia_por_id(db: Session, materia_id: int):
    return db.query(Materia).filter(Materia.id == materia_id).first()


def actualizar_materia(db: Session, materia_id: int, materia_data: MateriaUpdate):
    materia = db.query(Materia).filter(Materia.id == materia_id).first()
    if not materia:
        raise ValueError("Materia no encontrada")

    materia_data_dict = materia_data.model_dump(exclude_unset=True)

    for key, value in materia_data_dict.items():
        setattr(materia, key, value)

    db.commit()
    db.refresh(materia)

    return materia


def eliminar_materia(db: Session, materia_id: int):
    materia = db.query(Materia).filter(Materia.id == materia_id).first()
    if not materia:
        raise ValueError("Materia no encontrada")

    db.delete(materia)
    db.commit()
    return True


async def importar_materias(db: Session, byte_object: BytesIO):
    try:
        from src.infrastructure.database.orm_models import PlanEstudios, AreaConocimiento
        byte_object.seek(0)
        wb = xl.load_workbook(byte_object)
        sheet = wb["materias"]
        objects = []
        headers = [cell.value for cell in sheet[1]]

        # Construir mapeo de ID de Excel a Nombre del Plan de Estudios si existe la pestaña
        excel_plan_map = {}
        if "plan_estudios" in wb.sheetnames:
            plan_sheet = wb["plan_estudios"]
            plan_headers = [c.value for c in plan_sheet[1]]
            for r in plan_sheet.iter_rows(min_row=2, values_only=True):
                r_data = {k: v for k, v in zip(plan_headers, r)}
                ex_id = r_data.get("id")
                nombre = r_data.get("nombre")
                if ex_id is not None:
                    excel_plan_map[int(ex_id)] = nombre #type: ignore

        # Construir mapeo de ID de Excel a Nombre del Área de Conocimiento si existe la pestaña
        excel_area_map = {}
        if "areas_conocimiento" in wb.sheetnames:
            area_sheet = wb["areas_conocimiento"]
            area_headers = [c.value for c in area_sheet[1]]
            for r in area_sheet.iter_rows(min_row=2, values_only=True):
                r_data = {k: v for k, v in zip(area_headers, r)}
                ex_id = r_data.get("id")
                nombre = r_data.get("nombre")
                if ex_id is not None:
                    excel_area_map[int(ex_id)] = nombre #type: ignore

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {key: value for key, value in zip(headers, row)}
            if not any(value is not None and str(value).strip() != "" for value in row_data.values()):
                continue

            plan_id = None
            plan_id_raw = row_data.get("plan_estudios_id")
            if plan_id_raw is not None and str(plan_id_raw).strip() != "":
                # 1. Intentar resolver como entero (ID relativo Excel o ID BD)
                try:
                    ex_id = int(plan_id_raw) #type: ignore
                    if ex_id in excel_plan_map:
                        plan_name = excel_plan_map[ex_id]
                        plan_obj = db.query(PlanEstudios).filter(PlanEstudios.nombre == plan_name).first()
                        if plan_obj:
                            plan_id = plan_obj.id
                    if not plan_id: #type: ignore
                        exists = db.query(PlanEstudios).filter(PlanEstudios.id == ex_id).first()
                        if exists:
                            plan_id = exists.id
                except ValueError:
                    # Es una cadena con el nombre del plan directamente
                    plan_name = str(plan_id_raw).strip()
                    plan_obj = db.query(PlanEstudios).filter(PlanEstudios.nombre == plan_name).first()
                    if plan_obj:
                        plan_id = plan_obj.id
            
            # 2. Si no se resolvió, buscar por campos alternativos de texto
            if not plan_id: #type: ignore
                plan_name = str(row_data.get("plan_estudios_nombre") or row_data.get("plan_estudios") or row_data.get("plan") or "").strip()
                if plan_name:
                    plan_obj = db.query(PlanEstudios).filter(PlanEstudios.nombre == plan_name).first()
                    if plan_obj:
                        plan_id = plan_obj.id

            if not plan_id: #type: ignore
                raise ValueError(f"No se pudo resolver el Plan de Estudios para la materia: '{row_data.get('nombre_asignatura')}'")

            area_id = None
            area_id_raw = row_data.get("area_conocimiento_id") or row_data.get("area_conocimiento")
            if area_id_raw is not None and str(area_id_raw).strip() != "":
                # 1. Intentar resolver como entero (ID relativo Excel o ID BD)
                try:
                    ex_id = int(area_id_raw) #type: ignore
                    if ex_id in excel_area_map:
                        area_name = excel_area_map[ex_id]
                        area_obj = db.query(AreaConocimiento).filter(AreaConocimiento.nombre == area_name).first()
                        if area_obj:
                            area_id = area_obj.id
                    if not area_id: #type: ignore
                        exists = db.query(AreaConocimiento).filter(AreaConocimiento.id == ex_id).first()
                        if exists:
                            area_id = exists.id
                except ValueError:
                    # Es una cadena con el nombre del área directamente
                    area_name = str(area_id_raw).strip()
                    area_obj = db.query(AreaConocimiento).filter(AreaConocimiento.nombre == area_name).first()
                    if area_obj:
                        area_id = area_obj.id
            
            # 2. Si no se resolvió, buscar por campos alternativos de texto
            if not area_id: #type: ignore
                area_name = str(row_data.get("area_conocimiento_nombre") or row_data.get("area") or "").strip()
                if area_name:
                    area_obj = db.query(AreaConocimiento).filter(AreaConocimiento.nombre == area_name).first()
                    if area_obj:
                        area_id = area_obj.id

            if not area_id: #type: ignore
                raise ValueError(f"No se pudo resolver el Área de Conocimiento para la materia: '{row_data.get('nombre_asignatura')}'")

            materia_data = MateriaCreate(
                nombre_asignatura=str(row_data.get("nombre_asignatura") or "").strip(),
                plan_estudios_id=plan_id, #type: ignore
                numero_periodo=int(row_data.get("numero_periodo") or 0), #type: ignore
                hsm=int(row_data.get("hsm") or 0), #type: ignore
                area_conocimiento_id=area_id, #type: ignore
                estatus=str(row_data.get("estatus") or "ACTIVA"),
            )
            objects.append(crear_nueva_materia(db, materia_data))

        return objects
    except Exception as e:
        raise ValueError(f"Error al importar las materias: {str(e)}")


async def exportar_materias(materias: list):
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "materias" #type: ignore

    headers = ["id", "nombre_asignatura", "plan_estudios_id", "numero_periodo", "hsm", "area_conocimiento_id", "estatus"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header) #type: ignore

    for row_idx, materia in enumerate(materias, start=2):
        sheet.cell(row=row_idx, column=1, value=materia.id) #type: ignore
        sheet.cell(row=row_idx, column=2, value=materia.nombre_asignatura) #type: ignore
        sheet.cell(row=row_idx, column=3, value=materia.plan_estudios_id) #type: ignore
        sheet.cell(row=row_idx, column=4, value=materia.numero_periodo) #type: ignore
        sheet.cell(row=row_idx, column=5, value=materia.hsm) #type: ignore
        sheet.cell(row=row_idx, column=6, value=materia.area_conocimiento_id) #type: ignore
        sheet.cell(row=row_idx, column=7, value=getattr(materia.estatus, "value", materia.estatus)) #type: ignore

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer