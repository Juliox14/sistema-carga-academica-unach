from sqlalchemy.orm import Session
from src.infrastructure.database.orm_models import Docente, AreaConocimiento, EstatusDocente, DocenteUnidad
from src.infrastructure.api.schemas.docentes_schema import DocenteCreate, DocenteUpdate

def crear_docente(db: Session, docente_data: DocenteCreate, unidad_id: int | None = None):
    # 1. Separamos los IDs de las áreas del resto de los datos
    datos_dict = docente_data.model_dump(exclude={"areas_conocimiento_ids", "horas_obligatorias", "es_unidad_principal"})
    areas_ids = docente_data.areas_conocimiento_ids
    horas_obligatorias = docente_data.horas_obligatorias
    es_unidad_principal = docente_data.es_unidad_principal
    
    # Asegurar mayúsculas en nombre, apellidos y plaza
    if "nombre" in datos_dict and datos_dict["nombre"]:
        datos_dict["nombre"] = datos_dict["nombre"].upper()
    if "apellidos" in datos_dict and datos_dict["apellidos"]:
        datos_dict["apellidos"] = datos_dict["apellidos"].upper()
    if "plaza" in datos_dict and datos_dict["plaza"]:
        datos_dict["plaza"] = datos_dict["plaza"].upper()
    
    # 2. Creamos la instancia del docente
    nuevo_docente = Docente(**datos_dict)
    
    # 3. Si mandaron áreas de conocimiento, las buscamos y las vinculamos
    if areas_ids:
        # Hacemos un SELECT de las áreas cuyos IDs estén en la lista
        areas = db.query(AreaConocimiento).filter(AreaConocimiento.id.in_(areas_ids)).all()
        nuevo_docente.areas_conocimiento = areas
        
    # 4. Guardamos todo (SQLAlchemy inserta en docentes y en la tabla intermedia automáticamente)
    db.add(nuevo_docente)
    db.commit()
    db.refresh(nuevo_docente)
    # Vincular a la unidad academica si se provee
    if unidad_id:
        if es_unidad_principal:
            # Desmarcar otras unidades principales del docente si esta es la principal
            db.query(DocenteUnidad).filter(DocenteUnidad.docente_id == nuevo_docente.id).update({"es_unidad_principal": False})
        
        db.add(DocenteUnidad(
            docente_id=nuevo_docente.id, 
            unidad_academica_id=unidad_id,
            es_unidad_principal=es_unidad_principal,
            horas_obligatorias=horas_obligatorias
        ))
        db.commit()
        db.refresh(nuevo_docente)
    
    return nuevo_docente

def obtener_docentes(db: Session, unidad_id: int | None = None):
    q = db.query(Docente)
    if unidad_id is not None:
        q = q.join(DocenteUnidad, DocenteUnidad.docente_id == Docente.id).filter(
            DocenteUnidad.unidad_academica_id == unidad_id
        )
    return q.all()

def obtener_docente_por_id(db: Session, docente_id: int):
    return db.query(Docente).filter(Docente.id == docente_id).first()

def actualizar_docente(db: Session, docente_id: int, docente_data: DocenteUpdate, unidad_id: int | None = None):
    db_docente = db.query(Docente).filter(Docente.id == docente_id).first()
    if not db_docente:
        return None
        
    datos_actualizar = docente_data.model_dump(exclude_unset=True)
    
    horas_obligatorias = datos_actualizar.pop("horas_obligatorias", None)
    es_unidad_principal = datos_actualizar.pop("es_unidad_principal", None)
    
    # Asegurar mayúsculas en nombre, apellidos y plaza
    if "nombre" in datos_actualizar and datos_actualizar["nombre"]:
        datos_actualizar["nombre"] = datos_actualizar["nombre"].upper()
    if "apellidos" in datos_actualizar and datos_actualizar["apellidos"]:
        datos_actualizar["apellidos"] = datos_actualizar["apellidos"].upper()
    if "plaza" in datos_actualizar and datos_actualizar["plaza"]:
        datos_actualizar["plaza"] = datos_actualizar["plaza"].upper()
    
    if "areas_conocimiento_ids" in datos_actualizar:
        nuevos_ids = datos_actualizar.pop("areas_conocimiento_ids")
        nuevas_areas = db.query(AreaConocimiento).filter(AreaConocimiento.id.in_(nuevos_ids)).all()
        db_docente.areas_conocimiento = nuevas_areas
        
    for clave, valor in datos_actualizar.items():
        setattr(db_docente, clave, valor)
        
    db.commit()
    db.refresh(db_docente)
    
    # Actualizar vinculo con unidad si aplica
    if unidad_id and (horas_obligatorias is not None or es_unidad_principal is not None):
        vinculo = db.query(DocenteUnidad).filter_by(docente_id=docente_id, unidad_academica_id=unidad_id).first()
        if vinculo:
            if horas_obligatorias is not None:
                vinculo.horas_obligatorias = horas_obligatorias
            if es_unidad_principal is not None:
                if es_unidad_principal:
                    # Desmarcar otras unidades principales
                    db.query(DocenteUnidad).filter(
                        DocenteUnidad.docente_id == docente_id, 
                        DocenteUnidad.unidad_academica_id != unidad_id
                    ).update({"es_unidad_principal": False})
                vinculo.es_unidad_principal = es_unidad_principal
                
            db.commit()
            db.refresh(db_docente)
            
    return db_docente

def eliminar_docente(db: Session, docente_id: int):
    from src.infrastructure.database.orm_models import Usuario, PreferenciaDocente, OficioDocente, AsignacionOtraActividad, AsignacionCarga
    
    db_docente = db.query(Docente).filter(Docente.id == docente_id).first()
    if not db_docente:
        return False
    
    usuario_id = db_docente.usuario_id
    
    # 1. Eliminar preferencias vinculadas
    db.query(PreferenciaDocente).filter(PreferenciaDocente.docente_id == docente_id).delete()
    
    # 2. Eliminar oficios vinculados
    db.query(OficioDocente).filter(OficioDocente.docente_id == docente_id).delete()
    
    # 3. Eliminar asignaciones de otras actividades
    db.query(AsignacionOtraActividad).filter(AsignacionOtraActividad.docente_id == docente_id).delete()
    
    # 4. Desvincular asignaciones de carga docente
    db.query(AsignacionCarga).filter(AsignacionCarga.docente_titular_id == docente_id).update({AsignacionCarga.docente_titular_id: None})
    db.query(AsignacionCarga).filter(AsignacionCarga.docente_temporal_id == docente_id).update({AsignacionCarga.docente_temporal_id: None})
    
    # 5. Eliminar el docente
    db.delete(db_docente)
    
    # 6. Si tiene un usuario vinculado, eliminarlo para evitar usuarios huérfanos
    if usuario_id: #type: ignore
        usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
        if usuario:
            db.delete(usuario)
            
    db.commit()
    return True

async def importar_docentes(db: Session, byte_object, unidad_id: int | None = None):
    from io import BytesIO
    import openpyxl as xl
    from src.infrastructure.database.orm_models import CategoriaDocente, Docente, Turno
    try:
        byte_object.seek(0)
        wb = xl.load_workbook(byte_object)
        sheet = wb["docentes"] if "docentes" in wb.sheetnames else wb.active
        
        if sheet is None:
            raise ValueError("No se pudo encontrar la hoja de cálculo 'docentes' para importar.")
        objects = []
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {key: value for key, value in zip(headers, row) if key}
            if not any(value is not None and str(value).strip() != "" for value in row_data.values()):
                continue

            # 1. Obtener Categoría
            cat_val = str(row_data.get("Categoría") or row_data.get("Categoria") or row_data.get("categoria") or "").strip()
            categoria = None
            if cat_val:
                categoria = db.query(CategoriaDocente).filter(CategoriaDocente.siglas == cat_val).first()
                if not categoria:
                    categoria = db.query(CategoriaDocente).filter(CategoriaDocente.nombre == cat_val).first()
                if not categoria:
                    raise ValueError(f"La categoría '{cat_val}' del docente con plaza '{row_data.get('Plaza') or row_data.get('plaza') or 'SIN PLAZA'}' no existe en el sistema. Verifique las siglas.")
            else:
                raise ValueError(f"La categoría es obligatoria para el docente con plaza '{row_data.get('Plaza') or row_data.get('plaza') or 'SIN PLAZA'}'")
            
            categoria_id = categoria.id if categoria else 1

            # 2. Plaza
            plaza_raw = row_data.get("Plaza") or row_data.get("plaza") or ""
            if isinstance(plaza_raw, float):
                plaza = str(int(plaza_raw))
            else:
                plaza = str(plaza_raw).strip().upper()

            # 3. Nombres y Apellidos
            nombres = str(row_data.get("Nombres") or row_data.get("nombres") or row_data.get("Nombre") or "").strip().upper()
            apellidos = str(row_data.get("Apellidos") or row_data.get("apellidos") or "").strip().upper()

            # 4. Correo
            correo = row_data.get("Correo institucional") or row_data.get("Correo") or row_data.get("correo") or row_data.get("correo_institucional")
            correo = str(correo).strip() if correo is not None else None

            # 5. Teléfono
            tel = row_data.get("Teléfono") or row_data.get("Telefono") or row_data.get("telefono") or row_data.get("tel")
            tel = str(tel).strip() if tel is not None else None

            # 6. Estatus
            estatus_raw = str(row_data.get("Estatus") or row_data.get("estatus") or "ACTIVO").strip().upper()
            if estatus_raw in ("SABÁTICO", "SABATICO", "AÑO SABÁTICO", "AÑO SABATICO"):
                estatus = "SABATICO"
            elif estatus_raw in ("INACTIVO", "INACTIVA"):
                estatus = "INACTIVO"
            elif estatus_raw in ("ACTIVO", "ACTIVA"):
                estatus = "ACTIVO"
            elif estatus_raw in ("PERMISO",):
                estatus = "PERMISO"
            else:
                estatus = estatus_raw

            # 7. Turno
            turno_raw = str(row_data.get("Turno") or row_data.get("turno") or "MIXTO").strip().upper()
            if "MATUTINO" in turno_raw:
                turno = Turno.MATUTINO
            elif "VESPERTINO" in turno_raw:
                turno = Turno.VESPERTINO
            else:
                turno = Turno.MIXTO

            # 8. HSM Personalizadas
            hsm_raw = row_data.get("HSM personalizadas") or row_data.get("HSM") or row_data.get("hsm_personalizadas")
            hsm = None
            if hsm_raw is not None:
                try:
                    hsm_str = str(hsm_raw).strip()
                    if hsm_str not in ("", "—", "-", "_"):
                        hsm = float(hsm_raw) #type: ignore
                except ValueError:
                    pass

            if not nombres or not apellidos:
                continue

            status_str = estatus.strip() if estatus else "ACTIVO"
            if status_str == "SABATICO":
                status_str = "SABÁTICO" # Match seed with accent
            db_status = db.query(EstatusDocente).filter(EstatusDocente.nombre == status_str).first()
            if not db_status:
                db_status = EstatusDocente(nombre=status_str, permite_carga=True, es_prioritario=False)
                db.add(db_status)
                db.commit()
                db.refresh(db_status)
            estatus_id = db_status.id

            docente_db = None
            if plaza:
                docente_db = db.query(Docente).filter(Docente.plaza == plaza).first()

            if docente_db:
                docente_db.nombre = nombres #type: ignore
                docente_db.apellidos = apellidos #type: ignore
                docente_db.categoria_id = categoria_id #type: ignore
                docente_db.correo_institucional = correo #type: ignore
                docente_db.telefono = tel #type: ignore
                docente_db.estatus_id = estatus_id #type: ignore
                docente_db.turno = turno #type: ignore
                docente_db.hsm_personalizadas = hsm #type: ignore
                objects.append(docente_db)
            else:
                nuevo_docente = Docente(
                    nombre=nombres,
                    apellidos=apellidos,
                    plaza=plaza,
                    categoria_id=categoria_id,
                    correo_institucional=correo,
                    telefono=tel,
                    estatus_id=estatus_id,
                    turno=turno,
                    hsm_personalizadas=hsm
                )
                db.add(nuevo_docente)
                objects.append(nuevo_docente)

        db.commit()
        
        # Vincular a la unidad si aplica
        if unidad_id:
            for obj in objects:
                # Verificar si ya esta vinculado
                vinculo = db.query(DocenteUnidad).filter_by(docente_id=obj.id, unidad_academica_id=unidad_id).first()
                if not vinculo:
                    db.add(DocenteUnidad(docente_id=obj.id, unidad_academica_id=unidad_id))
            db.commit()
            
        for obj in objects:
            db.refresh(obj)

        return objects
    except Exception as e:
        db.rollback()
        raise ValueError(f"Error al importar los docentes: {str(e)}")